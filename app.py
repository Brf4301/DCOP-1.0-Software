import os
from flask import Flask, jsonify, render_template, request
from urllib.parse import unquote
from openpyxl import load_workbook
from torch_geometric.data import Data
from rdkit import Chem
from rdkit.Chem import Draw
import cv2
from PIL import Image
import pandas as pd
import torch
import numpy as np
from torch.optim.lr_scheduler import ReduceLROnPlateau
from torch_geometric.nn import GATConv, global_mean_pool
import torch.nn.functional as F
import torch.nn as nn
from pathlib import Path
import threading
import webbrowser
import waitress
import time
import shutil
import sys


# 动态设置静态文件和模板路径
if getattr(sys, 'frozen', False):
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
else:
    static_folder = os.path.join(Path(__file__).parent, 'static')
    app = Flask(__name__)

# print('static_folder: ', static_folder)

# 定义持久化数据目录  用于存放用户上传过来的数据
if getattr(sys, 'frozen', False):
    print("Running in a frozen environment")
    # Windows 示例  上传文件路径--本Windows系统的用户路径/appdata/Roaming/upload_data
    upload_dir = Path(os.environ['APPDATA']) / 'upload_data'
    database_dir = Path(os.environ['APPDATA']) / 'dataset_data'
    predict_dir = Path(os.environ['APPDATA']) / 'predict_data'  # 预测数据保存路径（上传后保存）
    database_dir.mkdir(exist_ok=True)
    # 将static_folder/dateset_data路径下的文件复制到database_dir
    if not os.listdir(database_dir):
        database_data_dir = os.path.join(static_folder, 'dataset_data')
        for fl in os.listdir(database_data_dir):
            if fl.endswith('.txt'):
                # print('copying file: ', fl)
                shutil.copy(os.path.join(database_data_dir, fl), database_dir)
else:
    print("Running in a non-frozen environment")
    upload_dir = Path.home() / 'upload_data'  # 本Windows系统的用户路径
    database_dir = Path('static') / 'dataset_data'
    predict_dir = Path.home() / 'predict_data'  # 预测数据保存路径（上传后保存）
# print(database_dir)
upload_dir.mkdir(exist_ok=True)
predict_dir.mkdir(exist_ok=True)


ALLOWED_EXTENSIONS = {'xlsx'}

# 需要跳过的路径列表
SKIP_PATHS = ['/upload', '/improt_newDataset']

short_full_name = {'CID': 'Chemical Identifier', 'Cmpdname': 'Compound Name', 'Em': 'ExactMass', 'Mf': 'Molecular Formula', 'Mw': 'Molecular Weight',
                   'Cmpdsynonym': 'Compound Synonyms', 'Inchi': 'International Chemical Identifier', 'Isosmiles': 'Isomeric SMILES', 'Iupacname': 'IUPAC Name'}
# 八大类别
class_items = ['Organophosphate Triester', 'Organophosphate Diester', 'Organophosphate Monoester',
               'Polyphosphates', 'Organophosphite', 'Organohypophosphite', 'Organophosphonate', 'Phosphine Oxide']

the_1_title = ['CID', 'Cmpdname', 'Exactmass', 'Mf', 'Mw', 'Cmpdsynonym', 'Inchi', 'Isosmiles', 'Iupacname']
the_3_fullTitle = ['Chemical Identifier', 'Name', 'Molecular Formula', 'Molecular Weight', 'Cas Number', 'Total tonnage Band', 'Last Updated', 'Registration Status',
                   'Registration Type', 'Isomeric SMILES', 'Inchikey', 'Exact Mass', 'Compound Synonyms', 'Topological Polar Surface Area', 'Complexity', 'XLogP',
                   'Heavy Atom Count', 'Hydrogen Bond Donor Count', 'Hydrogen Bond Acceptor Count', 'Rotatable Bond Count', 'International Chemical Identifier',
                   'IUPAC Name', 'Annotation Hits', 'BioAssay Identifiers']
the_3_title = ['Chemical Identifier', 'Name', 'Molecular Formula', 'Molecular Weight', 'Cas Number', 'Total tonnage Band', 'Last Updated', 'Registration Status',
               'Registration Type']
the_3_last_title = the_3_fullTitle[9:]
the_4_title = ['Number', 'Compound name', 'Abbr.', 'Chemical Formula', 'CAS Number', 'LogKOW', 'Adduct', 'Retention index', 'Exact Precursor mass',
               'Accurate fragment 1', 'Accurate fragment 2', 'Accurate fragment 3', 'Accurate fragment 4']
category_data_01 = {}
category_data_02 = {}
category_index = {}


# 定义模型
class MPNNWithAttention(nn.Module):
    def __init__(self, in_feats, h_feats=64, out_feats=326000, num_layers=4):
        super(MPNNWithAttention, self).__init__()
        self.num_layers = num_layers
        self.node_embedding = nn.Linear(in_feats, h_feats)
        self.mpnn_layers = nn.ModuleList(
            [GATConv(h_feats, h_feats, heads=4, concat=False),
             GATConv(h_feats, h_feats, heads=4, concat=False),
             GATConv(h_feats, h_feats, heads=4, concat=False)]
        )
        self.fc = nn.Linear(h_feats, out_feats)
        self.dropout = nn.Dropout(0.2)

    def forward(self, x, edge_index, edge_attr, batch):
        h = self.node_embedding(x)
        for layer in self.mpnn_layers:
            # print(h.size())
            h = F.relu(layer(h, edge_index))
        h = global_mean_pool(h, batch)
        h = self.dropout(h)
        return self.fc(h)


def open_browser():
    # 等待 1 秒确保服务器启动完成
    time.sleep(1)
    webbrowser.open_new('http://127.0.0.1:5000/')


# 选择设备，GPU优先
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


def sanitize_filename(filename):
    """安全处理文件名"""
    # 解码URL编码字符（处理中文文件名）
    filename = unquote(filename)
    # 去除路径信息
    filename = os.path.basename(filename)
    # 替换系统保留字符
    replace_chars = {'\\', '/', ':', '*', '?', '"', '<', '>', '|'}
    for char in replace_chars:
        filename = filename.replace(char, '_')
    return filename.strip()


def allowed_file(filename):
    """验证文件扩展名是否合法"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def save_data_to_txt(folder1, folder2):
    for fl in os.listdir(folder1):
        if not fl.endswith('.xlsx'):
            continue
        file_path = os.path.join(folder1, fl)
        # 创建 ExcelFile 对象
        excel_file = pd.ExcelFile(file_path)

        # 获取所有 sheet 名称的列表
        sheet_names = excel_file.sheet_names
        excel_file.close()
        if 'Organophosphite' in sheet_names:
            workbook = load_workbook(file_path, read_only=True)
            for item in class_items:
                sheet = workbook[item]
                save_txt_path = os.path.join(folder2, item + '.txt')
                i = 0
                with open(save_txt_path, 'w', encoding='utf-8') as f:
                    for row in sheet.iter_rows(min_col=1, max_col=9, values_only=True):
                        if i == 0:
                            print(row)
                            i += 1
                        else:
                            f.write(str(row) + '\n')
            workbook.close()
            print('第一 第二部分： 类别', fl)

        elif 'HPVC-OPCs' in sheet_names or 'HPVC-OPCs'.lower() in sheet_names:
            workbook = load_workbook(file_path, read_only=True)
            try:
                sheet = workbook['HPVC-OPCs']
            except Exception:
                sheet = workbook['HPVC-OPCs'.lower()]
            save_txt_path = os.path.join(folder2, 'HPVC-OPCs.txt')
            i = 0
            with open(save_txt_path, 'w', encoding='utf-8') as f:
                for row in sheet.iter_rows(min_col=1, max_col=24, values_only=True):
                    if i == 0:
                        i += 1
                    else:
                        f.write(str(row) + '\n')
            workbook.close()
            print('第三部分： HPVC-OPCs', fl)
        elif 'OPCs质谱碎片信息' in sheet_names:
            workbook = load_workbook(file_path, read_only=True)
            sheet = workbook['OPCs质谱碎片信息']
            save_txt_path = os.path.join(folder2, 'OPCs质谱碎片信息.txt')
            i = 0
            with open(save_txt_path, 'w', encoding='utf-8') as f:
                for row in sheet.iter_rows(min_col=1, max_col=13, values_only=True):
                    if i == 0 or i == 1:
                        i += 1
                    else:
                        f.write(str(row) + '\n')
            workbook.close()
            print('第四部分： OPCs质谱碎片信息', fl)


def query_data(key_val, index):
    data = ()
    category = ''
    have_found = False
    for item in class_items:
        data_txt_path = os.path.join(database_dir, item + '.txt')
        if os.path.exists(data_txt_path):
            with open(data_txt_path, 'r', encoding='utf-8') as f:
                for line in f:
                    tup_data = eval(line.strip())
                    if str(tup_data[index]) == key_val:
                        data = tup_data
                        category = item
                        have_found = True
                        break
            if have_found:
                break
    return data, category


def get_category_count(class_item_data):
    category_count_dict = {}
    for item in class_item_data:
        data_txt_path = os.path.join(database_dir, item + '.txt')
        if os.path.exists(data_txt_path):
            with open(data_txt_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                count = 0
                for one in lines[-3:]:
                    if not eval(one.strip())[0]:
                        count += 1
                category_count_dict[item] = len(lines)-count
    return category_count_dict


def get_loc_page_index(N, page_size, page):
    if N/page_size > N//page_size:
        total_pages = N//page_size + 1
    else:
        total_pages = N//page_size

    if page > total_pages:
        page = total_pages
    if page < 1:
        page = 1

    start_index = (page-1)*page_size
    if page < total_pages:
        end_index = page*page_size
    else:
        end_index = N

    return [start_index, end_index, total_pages]


@app.before_request
def check_dataset():
    upload_dir.mkdir(exist_ok=True)  # 自动创建空文件夹
    database_dir.mkdir(exist_ok=True)
    if not os.listdir(database_dir):
        if os.listdir(upload_dir):
            print('正在将数据集excel文件数据导入到本地数据库文件...')
            save_data_to_txt(upload_dir, database_dir)
        else:
            if request.path in SKIP_PATHS:
                return  # 直接跳过
            return render_template('home0.html')


# 定义 mol_to_graph 函数
def mol_to_graph(mol):
    node_features = []
    for atom in mol.GetAtoms():
        features = [0] * 12  # 新的元素集合：C, N, O, F, S, Cl, Na, Al, P, Br, K, Mg
        atomic_num = atom.GetAtomicNum()
        # 根据元素的原子序号设置对应特征
        if atomic_num == 6:  # C
            features[0] = 1
        elif atomic_num == 7:  # N
            features[1] = 1
        elif atomic_num == 8:  # O
            features[2] = 1
        elif atomic_num == 9:  # F
            features[3] = 1
        elif atomic_num == 16:  # S
            features[4] = 1
        elif atomic_num == 17:  # Cl
            features[5] = 1
        elif atomic_num == 11:  # Na
            features[6] = 1
        elif atomic_num == 13:  # Al
            features[7] = 1
        elif atomic_num == 15:  # P
            features[8] = 1
        elif atomic_num == 35:  # Br
            features[9] = 1
        elif atomic_num == 19:  # K
            features[10] = 1
        elif atomic_num == 12:  # Mg
            features[11] = 1
        node_features.append(features)

    # 构建边特征
    edge_indices = []
    edge_attr = []  # 初始化 edge_attr

    for bond in mol.GetBonds():
        bond_type = bond.GetBondTypeAsDouble()
        bond_features = [0] * 6  # 单键/双键/三键/芳香键/环状键/其他
        if bond_type == 1.0:
            bond_features[0] = 1
        elif bond_type == 2.0:
            bond_features[1] = 1
        elif bond_type == 3.0:
            bond_features[2] = 1
        if bond.GetIsAromatic():
            bond_features[3] = 1
        if bond.IsInRing():
            bond_features[4] = 1
        edge_indices.append([bond.GetBeginAtomIdx(), bond.GetEndAtomIdx()])
        edge_indices.append([bond.GetEndAtomIdx(), bond.GetBeginAtomIdx()])
        edge_attr.append(bond_features)
        edge_attr.append(bond_features)  # 无向图需要双向边

    # 转换 edge_index 和 edge_attr 为 PyTorch 张量
    edge_index = torch.tensor(edge_indices, dtype=torch.long).t().contiguous()
    edge_attr = torch.tensor(edge_attr, dtype=torch.float)

    # 转换节点特征为 PyTorch 张量
    node_features = torch.tensor(node_features, dtype=torch.float)

    return edge_index, node_features, edge_attr


@app.route('/')
def home():
    if os.listdir(database_dir):
        data = []
        for item in class_items:
            data_txt_path = os.path.join(database_dir, item + '.txt')
            if os.path.exists(data_txt_path):
                with open(data_txt_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        data.append(eval(line.strip()))
                        if len(data) > 10:
                            break
                if len(data) > 10:
                    break
        return render_template('home.html', data=data, title=the_1_title)
    else:
        if os.listdir(upload_dir):
            save_data_to_txt(upload_dir, database_dir)
            if os.listdir(database_dir):
                data = []
                for item in class_items:
                    data_txt_path = os.path.join(database_dir, item + '.txt')
                    if os.path.exists(data_txt_path):
                        with open(data_txt_path, 'r', encoding='utf-8') as f:
                            for line in f:
                                data.append(eval(line.strip()))
                                if len(data) > 10:
                                    break
                        if len(data) > 10:
                            break
                return render_template('home.html', data=data, title=the_1_title)
            else:
                data = []
                print('上传的文件数据有误！请检查上传的文件是否是目标数据')
                for item in class_items:
                    data_txt_path = os.path.join(database_dir, item + '.txt')
                    if os.path.exists(data_txt_path):
                        with open(data_txt_path, 'r', encoding='utf-8') as f:
                            for line in f:
                                print(line)
                        break

                return render_template('home0.html')
        else:
            return render_template('home0.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    if 'files' not in request.files:
        print('No files selected')
        return jsonify(success=False, message="No files selected")

    files = request.files.getlist('files')
    results = {'success': [], 'errors': []}

    for file in files:
        if file.filename == '':
            results['errors'].append("Empty filename detected")
            continue

        # 安全处理文件名（保留原始名称但清理危险字符）
        raw_filename = file.filename
        safe_filename = sanitize_filename(raw_filename)

        # 验证文件名有效性
        if safe_filename == '':
            results['errors'].append(f"Invalid file name：{raw_filename}")
            continue

        # 验证文件类型
        if not allowed_file(safe_filename):
            results['errors'].append(f"Unsupported file type: {raw_filename}")
            continue

        try:
            # 构建完整保存路径
            save_path = os.path.join(upload_dir, safe_filename)
            # 防止目录遍历攻击
            if not os.path.realpath(save_path).startswith(os.path.realpath(upload_dir)):
                raise ValueError("Illegal storage path")

            file.save(save_path)
            results['success'].append({
                'original': raw_filename,
                'saved': safe_filename
            })
        except Exception as e:
            results['errors'].append(f"{raw_filename} Save failed：{str(e)}")

    return jsonify(
        success=bool(results['success']),
        message=f"Successfully uploaded {len(results['success'])} files，failed {len(results['errors'])} files.",
        details=results
    )


@app.route('/key_search', methods=['GET', 'POST'])
def key_search():
    if request.method == 'GET':
        return render_template("key_search.html")
    key = request.form.get('key')  # 提取搜索内容 关键信息
    if key == '':
        return render_template("key_search.html")

    key_choose = request.form.get("choose")
    if key_choose == '1':
        print('CID_search')
        data, category = query_data(key, 0)
    elif key_choose == '2':
        print('Compound Name search')
        data, category = query_data(key, 1)
    else:
        print('Isomeric SMILES search')
        data, category = query_data(key, 7)

    if data:
        smiles = data[7]
        # 根据smiles生成化学结构图
        mol = Chem.MolFromSmiles(smiles)
        if mol is None:
            not_found_s = "Invalid SMILES!."
            return render_template("notFound.html", not_found_s=not_found_s, key=key, key_choose=key_choose)

        # 配置绘制选项
        draw_options = Draw.MolDrawOptions()
        draw_options.backgroundColour = (0.941, 0.973, 1.0)  # AliceBlue的归一化RGB值
        draw_options.bondLineWidth = 2  # 适当加粗键的线条
        draw_options.clearBackground = False  # 关键修改：禁止透明背景

        # 生成图像
        img = Draw.MolToImage(mol, size=(400, 300), options=draw_options, kekulize=True)

        # 手动覆盖背景（双重保障）
        background_color = (240, 248, 255)  # AliceBlue的RGB值
        if img.mode == 'RGBA':
            # 创建纯色背景
            bg = Image.new("RGB", img.size, background_color)
            # 合并图像
            bg.paste(img, mask=img.split()[3])  # 使用alpha通道作为掩码
            img = bg

        img_np = np.array(img)
        # 读取时指定灰度模式
        # gray_img = cv2.imread(img_np, cv2.IMREAD_GRAYSCALE)
        img_opencv = cv2.cvtColor(img_np, cv2.COLOR_RGB2BGR)

        # 保存图像
        img_path = os.path.join(static_folder, 'chemical_structure.png')
        cv2.imwrite(img_path, img_opencv)
        return render_template("search_result.html", key=key, category=category, title=the_1_title, data=data, key_choose=key_choose)
    else:
        not_found_s = "No relevant data was found."
        return render_template("notFound.html", not_found_s=not_found_s, key=key, key_choose=key_choose)


@app.route('/category')
def category_search():
    global category_data_01, category_data_02, category_index
    for i in range(len(class_items)):
        category_index[class_items[i]] = str(i+1)  # 每一类对应一个序号 用于对应前端页面对应元素id
    category_data_01 = get_category_count(class_items[:4])
    category_data_02 = get_category_count(class_items[4:])
    return render_template("category_summary.html", category_data_01=category_data_01, category_data_02=category_data_02)


@app.route('/category_result/<category>', methods=['GET', 'POST'])
def category_result(category):
    if request.method == 'GET':
        all_data = []
        page = request.args.get('page', 1, type=int)  # 获取当前页面参数
        page_size = request.args.get('page_size', 10, type=int)
        # 读取对应类别category数据库文件txt，得到所有数据data
        data_txt_path = os.path.join(database_dir, category + '.txt')
        if os.path.exists(data_txt_path):
            with open(data_txt_path, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.strip():
                        tup_data = eval(line.strip())
                        all_data.append(tup_data)
        print('数据总个数: ', len(all_data))
        loc_index_total = get_loc_page_index(len(all_data), page_size, page)  # 得到当前页的索引span，和总页数
        # 截取当前页要展示的数据
        data = all_data[loc_index_total[0]:loc_index_total[1]]
        total_page = loc_index_total[2]
        return render_template("category_result.html", category_data_01=category_data_01,
                               category_data_02=category_data_02, data=data, title=the_1_title,
                               category=category, category_index=category_index, page=page,
                               page_size=page_size, total_page=total_page, all_data=all_data)
    all_data = []
    page = int(request.form.get('cur_page'))
    page_size = int(request.form.get('page_size'))
    # 读取对应类别category数据库文件txt，得到所有数据data
    data_txt_path = os.path.join(database_dir, category + '.txt')
    if os.path.exists(data_txt_path):
        with open(data_txt_path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    tup_data = eval(line.strip())
                    all_data.append(tup_data)
    print('数据总个数: ', len(all_data))
    loc_index_total = get_loc_page_index(len(all_data), page_size, page)  # 得到当前页的索引span，和总页数
    # 截取当前页要展示的数据
    data = all_data[loc_index_total[0]:loc_index_total[1]]
    total_page = loc_index_total[2]
    return render_template("category_result.html", category_data_01=category_data_01,
                           category_data_02=category_data_02, data=data, title=the_1_title,
                           category=category, category_index=category_index,
                           page=page, page_size=page_size, total_page=total_page, all_data=all_data)


@app.route('/HPVC', methods=['GET', 'POST'])
def hpvc():
    page_size = 20  # 每页展示20条数据 写死了
    all_data = []
    if request.method == 'GET':
        page = request.args.get('page', 1, type=int)  # 获取当前页面参数
    else:
        page = int(request.form.get('cur_page'))
    # 读取对应类别category数据库文件txt，得到所有数据data
    data_txt_path = os.path.join(database_dir, 'HPVC-OPCs.txt')
    if os.path.exists(data_txt_path):
        with open(data_txt_path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    tup_data = eval(line.strip())
                    all_data.append(tup_data)
    total = len(all_data)
    loc_index_total = get_loc_page_index(len(all_data), page_size, page)  # 得到当前页的索引span，和总页数
    # 截取当前页要展示的数据
    data = all_data[loc_index_total[0]:loc_index_total[1]]
    total_page = loc_index_total[2]
    detail_data = {}
    for one in data:
        detail_data[str(one[0])] = one[9:]
    return render_template("HPVC.html", data=data, page=page, total_page=total_page, total=total,
                           title=the_3_title, other_title=the_3_last_title, detail_data=detail_data)


# 更新数据库 （将原来已有数据库文件全部删除，导入新数据）
@app.route('/improt_newDataset')
def improt_newDataset():
    # 删除旧的数据库文件，导入新的数据库文件
    # shutil.rmtree(upload_dir)
    for fl in os.listdir(upload_dir):  # 删除上传的文件
        os.remove(os.path.join(upload_dir, fl))
    for fl in os.listdir(database_dir):  # 删除数据库文件
        os.remove(os.path.join(database_dir, fl))
    # shutil.rmtree(database_dir)
    notice = "The old database has been deleted. Now you can import the new one."
    return render_template('home0.html', notice=notice)


@app.route('/ms_fragment', methods=['GET'])
def ms_fragment():
    data_txt_path = os.path.join(database_dir, 'OPCs质谱碎片信息.txt')
    data = []
    if os.path.exists(data_txt_path):
        with open(data_txt_path, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    tup_data = eval(line.strip())
                    data.append(tup_data)
    return render_template("ms_fragment.html", data=data, title=the_4_title)


@app.route('/prediction_index', methods=['GET'])
def prediction_index():
    return render_template('prediction_index.html')


@app.route('/prediction/choose', methods=['GET'])
def prediction_choose():
    choose = request.args.get('para')
    # 单个预测
    if choose == '1':
        return render_template('pre_single_in.html')
    else:  # 批量预测
        # 上传预测数据文件之前先将目标文件夹中已存在的预测文件删除
        for fl in os.listdir(predict_dir):
            os.remove(os.path.join(predict_dir, fl))
        return render_template("upload_pred_file3.html")


@app.route('/prediction/single', methods=['POST'])
def prediction_single():
    smiles = request.form.get('smiles')
    if smiles == '':
        error = 'Please enter SMILES'
        return render_template('pre_single_in.html', msg=error)

    # 转换为分子对象
    mol = Chem.MolFromSmiles(smiles)
    if not mol:
        print("Invalid SMILES ")
        return render_template('pre_single_in.html', msg="Invalid SMILES ")

    # 数据库查找该smiles的基本信息
    base_data, category = query_data(smiles, 7)
    if not base_data:
        print('没找到数据')
        not_found_s = "No relevant data was found."
        return render_template('pre_single_in.html', msg=not_found_s)

    Chem.Draw.MolToImage(mol)

    model = MPNNWithAttention(in_feats=12, h_feats=256, out_feats=326000, num_layers=3).to(device)

    optimizer = torch.optim.AdamW(model.parameters(), lr=1e-3, weight_decay=1e-4)
    scheduler = ReduceLROnPlateau(optimizer,
                                  mode='min',  # 'min' 表示目标是最小化验证损失
                                  factor=0.1,  # 学习率每次降低的比例，例如从 0.001 -> 0.0001
                                  patience=10,  # 容忍验证损失在 10 个 epoch 内不下降
                                  #   verbose=True,  # 输出学习率调整信息
                                  min_lr=1e-6)  # 最低学习率限制

    # 初始化变量
    # train_losses = []
    # val_losses = []
    # cosine_similarities = []
    # epoch = 0

    # 加载模型的状态
    mode_file = os.path.join(static_folder, "mpnn_with_base256_20+.pth")
    checkpoint = torch.load(mode_file,  map_location=device)  # 加载训练好的模型文件，路径需自己对应更改
    model.load_state_dict(checkpoint['model_state_dict'])
    optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
    scheduler.load_state_dict(checkpoint['scheduler_state_dict'])
    # train_losses = checkpoint['train_losses']
    # val_losses = checkpoint['val_losses']
    # epoch = checkpoint['epoch']
    # 修复 cosine_similarities 的问题
    cosine_similarities = checkpoint.get('mean_cosine_similarity', [])
    if isinstance(cosine_similarities, (float, np.float32)):  # 如果是标量
        cosine_similarities = [cosine_similarities]  # 转为列表
    elif not isinstance(cosine_similarities, list):
        raise ValueError("cosine_similarities 应该是一个列表或数组类型")

    print("模型已从 'mpnn_with_attention.pth' 加载")

    # 打印模型、优化器和调度器的状态
    print(model)
    print(optimizer)
    print(scheduler)
    print(f"当前调度器状态: {scheduler.state_dict()}")

    mz_range = (74, 400)  # 根据实际情况调整范围

    # 模型评估模式
    model.eval()

    # 模型预测逻辑修正
    # 模型预测逻辑修正
    with torch.no_grad():
        # 转换为图数据
        edge_index, node_features, edge_attr = mol_to_graph(mol)
        node_features = node_features.to(device)
        edge_index = edge_index.to(device)
        edge_attr = edge_attr.to(device)
        # 创建 Data 对象
        graph_data = Data(
            x=node_features,
            edge_index=edge_index,
            edge_attr=edge_attr
        )
        # 模型预测
        graph_data = graph_data.to(device)
        out = model(
            graph_data.x,
            graph_data.edge_index,
            graph_data.edge_attr,
            torch.zeros(graph_data.x.shape[0], dtype=torch.long, device=device)  # 假设 batch=0
        )
        # 检查输出形状
        print(f"Output shape: {out.shape}")
        predicted_spectrum = out.squeeze().cpu().numpy()  # 去掉 batch 维度
        # bins = 8192  # 调整为较小的分箱数326000
        bins = 326000  # 调整为较小的分箱数326000
        mz_values = np.linspace(mz_range[0], mz_range[1], bins)
        if len(predicted_spectrum) != bins:
            raise ValueError(f"预测谱长度 ({len(predicted_spectrum)}) 和频谱分箱数 ({bins}) 不匹配！")
        # 获取强度最高的前 200 个峰
        top_indices = np.argsort(predicted_spectrum)[-100:]
        filtered_mz_values = mz_values[top_indices]
        filtered_intensities = predicted_spectrum[top_indices]
        # 将 m/z 和 intensity 展平到行
        # result_row = {}
        mz_datas = []
        intensity_datas = []
        for i, (mz, intensity) in enumerate(zip(filtered_mz_values, filtered_intensities)):
            # result_row[f'mz_{i+1}'] = mz
            # result_row[f'intensity_{i+1}'] = intensity
            mz_datas.append(mz)
            intensity_datas.append(intensity)

        return render_template("prediction_single_res.html", base_data=base_data, mz_datas=mz_datas, intensity_datas=intensity_datas)


# 上传预测数据
@app.route('/upload_pred_file', methods=['POST'])
def upload_predict_file():
    if 'files' not in request.files:
        print('No files selected')
        return jsonify(success=False, message="No files selected")

    files = request.files.getlist('files')
    results = {'success': [], 'errors': []}

    for file in files:
        if file.filename == '':
            results['errors'].append("Empty filename detected")
            continue

        # 安全处理文件名（保留原始名称但清理危险字符）
        raw_filename = file.filename
        safe_filename = sanitize_filename(raw_filename)

        # 验证文件名有效性
        if safe_filename == '':
            results['errors'].append(f"Invalid file name：{raw_filename}")
            continue

        # 验证文件类型
        if not allowed_file(safe_filename):
            results['errors'].append(f"Unsupported file type: {raw_filename}")
            continue

        try:
            # 构建完整保存路径
            save_path = os.path.join(predict_dir, safe_filename)
            # 防止目录遍历攻击
            if not os.path.realpath(save_path).startswith(os.path.realpath(predict_dir)):
                raise ValueError("Illegal storage path")

            file.save(save_path)
            results['success'].append({
                'original': raw_filename,
                'saved': safe_filename
            })
        except Exception as e:
            results['errors'].append(f"{raw_filename} Save failed：{str(e)}")

    return jsonify(
        success=bool(results['success']),
        message=f"Successfully uploaded {len(results['success'])} files，failed {len(results['errors'])} files.",
        details=results
    )


@app.route('/upload_pred_file2', methods=['POST'])
def upload_file():
    # Check if the post request has the file part
    if 'files' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    files = request.files.getlist('files')
    results = []

    for file in files:
        # If the user does not select a file, the browser submits an
        # empty file without a filename.
        if file.filename == '':
            results.append({'filename': '', 'status': 'No selected file'})
            continue

        if file and allowed_file(file.filename):
            # Secure the filename to prevent directory traversal
            safe_filename = sanitize_filename(file.filename)
            file_path = os.path.join(predict_dir, safe_filename)
            file.save(file_path)
            results.append({'filename': safe_filename, 'status': 'success'})
        else:
            results.append({'filename': file.filename, 'status': 'Invalid file type'})

    return jsonify({'results': results}), 200


# 批量预测
@app.route('/predict/batch', methods=['GET'])
def predict_batch():
    pred_files = os.listdir(predict_dir)
    if pred_files:
        file_path = os.path.join(predict_dir, pred_files[0])

        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)

        # 检查是否有 'SMILES' 列
        if 'SMILES' not in df.columns:
            # 先删为敬
            for fl in os.listdir(predict_dir):
                os.remove(os.path.join(predict_dir, fl))
            return render_template("upload_pred_file3.html", msg="The input file must contain a 'SMILES' column.")

        # 初始化模型和优化器
        model = MPNNWithAttention(in_feats=12, h_feats=256, out_feats=326000, num_layers=3).to(device)
        optimizer = torch.optim.AdamW(model.parameters(), lr=1e-3, weight_decay=1e-4)
        scheduler = ReduceLROnPlateau(optimizer,
                                      mode='min',  # 'min' 表示目标是最小化验证损失
                                      factor=0.1,  # 学习率每次降低的比例，例如从 0.001 -> 0.0001
                                      patience=10,  # 容忍验证损失在 10 个 epoch 内不下降
                                      #   verbose=True,  # 输出学习率调整信息
                                      min_lr=1e-6)  # 最低学习率限制

        # 加载模型的状态
        mode_file = os.path.join(static_folder, "mpnn_with_base256_20+.pth")
        checkpoint = torch.load(mode_file,  map_location=device)  # 加载训练好的模型文件，路径需自己对应更改
        model.load_state_dict(checkpoint['model_state_dict'])
        optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
        scheduler.load_state_dict(checkpoint['scheduler_state_dict'])

        # 打印模型、优化器和调度器的状态
        # print(model)
        # print(optimizer)
        # print(scheduler)
        print(f"当前调度器状态: {scheduler.state_dict()}")

        mz_range = (74, 400)  # 根据实际情况调整范围
        # 预测结果存储
        results = []

        # 模型评估模式
        model.eval()

        # 模型预测逻辑修正
        with torch.no_grad():
            for index, row in df.iterrows():
                smiles = str(row['SMILES'])  # 确保 SMILES 是字符串
                compound_id = str(row.get('id', index))  # 确保 id 是字符串
                # 转换为分子对象
                mol = Chem.MolFromSmiles(smiles)
                if not mol:
                    print(f"Invalid SMILES for row {index}: {smiles}")
                    continue
                # 转换为图数据
                edge_index, node_features, edge_attr = mol_to_graph(mol)
                node_features = node_features.to(device)
                edge_index = edge_index.to(device)
                edge_attr = edge_attr.to(device)
                # 创建 Data 对象
                graph_data = Data(
                    x=node_features,
                    edge_index=edge_index,
                    edge_attr=edge_attr
                )
                # 模型预测
                graph_data = graph_data.to(device)
                out = model(
                    graph_data.x,
                    graph_data.edge_index,
                    graph_data.edge_attr,
                    torch.zeros(graph_data.x.shape[0], dtype=torch.long, device=device)  # 假设 batch=0
                )
                # 检查输出形状
                print(f"Output shape: {out.shape}")
                predicted_spectrum = out.squeeze().cpu().numpy()  # 去掉 batch 维度
                # bins = 8192  # 调整为较小的分箱数326000
                bins = 326000  # 调整为较小的分箱数326000
                mz_values = np.linspace(mz_range[0], mz_range[1], bins)
                if len(predicted_spectrum) != bins:
                    raise ValueError(f"预测谱长度 ({len(predicted_spectrum)}) 和频谱分箱数 ({bins}) 不匹配！")
                # 获取强度最高的前 200 个峰
                top_indices = np.argsort(predicted_spectrum)[-100:]
                filtered_mz_values = mz_values[top_indices]
                filtered_intensities = predicted_spectrum[top_indices]
                # 将 m/z 和 intensity 展平到行
                result_row = {'id': compound_id, 'SMILES': smiles}
                for i, (mz, intensity) in enumerate(zip(filtered_mz_values, filtered_intensities)):
                    result_row[f'mz_{i+1}'] = mz
                    result_row[f'intensity_{i+1}'] = intensity
                results.append(result_row)

        # 将结果转换为 DataFrame 并保存
        results_df = pd.DataFrame(results)
        results_df = results_df.astype(str)  # 确保所有列为字符串类型
        results_df.to_csv('result.csv', index=False)
        # 预测完成之后 删除预测文件
        for fl in os.listdir(predict_dir):
            os.remove(os.path.join(predict_dir, fl))

        return render_template("upload_pred_file3.html", success='The prediction has been completed.Predicted result have saved to result.csv ')
    else:
        return render_template("upload_pred_file3.html", msg="Please upload the prediction file!")


# introduction
@app.route('/introduction')
def help_fun():
    return render_template("introduction.html")


if __name__ == '__main__':
    # app.run(debug=False)
    # 在子线程中打开浏览器，避免阻塞主线程
    threading.Thread(target=open_browser).start()
    waitress.serve(app, host='127.0.0.1', port=5000)
